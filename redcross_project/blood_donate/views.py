from django.shortcuts import render,redirect,get_object_or_404
from .models import HomeNav_drop
from django.http import HttpResponseRedirect
import requests
from django.contrib import messages
import json
from .models import pre_registataion_QA_1,pre_registataion_QA_2,pre_registataion_QA_3,pre_registataion_QA_4
from .models import appfooter,appfooter1,appdown,gettouch,socialicon
from .models import blood,Announcement,events,aboutus , bloodtype
from .models import patient_request,basic_requirements , donators , additional_requirements,Donor,Address
from .models import contact,contactus,Contactusadmin,tips_donation,content,UserProfile
from django.http import HttpResponse
from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout,login,authenticate
from django.contrib.auth.models import User
from django.db.models import Q
# Create your views here.

@login_required
def admindashboard(request):
    return render(request,"admindashboard.html")


@login_required(login_url='/admin_login/')
def administrate(request):
    donor_count = Donor.objects.count()
    news_count=events.objects.count()
    return render(request,"card.html",{'donor_count':donor_count,'news_count':news_count})
    
def countdown(request):
    return render(request , "countdown.html")

def homepage(request):
    
   
    blood_images = blood.objects.all()
    announcements = Announcement.objects.all()
    
    return render(request,"index.html" ,{'blood_images':blood_images,'announcements':announcements})
def test(request):
    return render(request,"test.html")
def cards(request):
    return render(request,"flashdesign.html")

def search_blood(request):
   
    countries = Country.objects.all()
    donorlisters = Donor.objects.all()  
      # Fetch your donor list based on your criteria
    eligible_donors = [donor for donor in donorlisters if donor.has_filled_form() and donor.is_eligible()]

    # Filter donors who haven't filled out the form
    donors_without_form = [donor for donor in donorlisters if not donor.has_filled_form()]

    donorlist = eligible_donors + donors_without_form

      
    
    if request.method=="POST":
        
        blood_group = request.POST.get('blood', '')
        country_id = request.POST.get('country', '')
        
        state_id = request.POST.get('state', '')
        district_id = request.POST.get('district', '')
        if request.user.is_authenticated:
            
            if blood_group :
                donorlist = donorlisters.filter(bloodgroup=blood_group)
            

            if country_id:
                donorlist = donorlist.filter(address__country_id=country_id)
            if state_id:
                donorlist = donorlist.filter(address__state_id=state_id)

            if district_id:
                donorlist = donorlist.filter(address__district_id=district_id)

            

        
            # donorlist = eligible_donorlist
            
            # # Get donors who have filled the form and are eligible
            # eligible_donors = Donor.objects.filter(Q(has_filled_form=True) & Q(eligible_donorlist))

            # # Get donors who haven't filled the form
            # donors_without_form = Donor.objects.filter(has_filled_form=False)

            # # Get donors who have filled the form but are not eligible


            # donors_with_form = Donor.objects.filter(has_filled_form=True)

            # # Filter donors who are eligible after filling out the form using Python
            # eligible_after_filling_form = [donor for donor in donors_with_form if donor.is_eligible()]

            # donorlist = list(donors_without_form) + eligible_after_filling_form

            eligible_donors = [donor for donor in donorlist if donor.has_filled_form() and donor.is_eligible()]

    # Filter donors who have filled out the form but are not eligible
            ineligible_donors = [donor for donor in donorlist if donor.has_filled_form() and not donor.is_eligible()]

    # Filter donors who haven't filled out the form
            donors_without_form = [donor for donor in donorlist if not donor.has_filled_form()]

            donorlist = donors_without_form + eligible_donors


        else:
            messages.error(request, 'Please login to search for blood donors.')
            donorlist = []

        
    return render(request,"search_donar.html" ,{'donorlist':donorlist  , 'countries' :countries})




def about(request):
    if request.method=="GET":
        k=aboutus.objects.all()
        return render(request,"about.html",{'k':k})


# Create your views here.
def footer(request):
    a=appfooter.objects.all()
    a1=appfooter1.objects.all()
    a2=appdown.objects.all()
    a3=gettouch.objects.all()
    a4=socialicon.objects.all()
    return render(request,"appfooter.html",{'a':a,'a1':a1,'a2':a2,'a3':a3,'a4':a4})

from .models import events
def event(request):
    if request.method=="GET":
     k=events.objects.all()
    return render(request,"event.html",{'k':k})






# Create your views here.


def blood_carousel(request):
    blood_images = blood.objects.all()
    announcements = Announcement.objects.all()
    return render(request, 'Carousel/blood_carousel.html', {'blood_images':blood_images,'announcements':announcements})

def  blood_carousel_img_form(request):
    q=blood.objects.all()
    if request.method=="POST":
        photo=request.FILES.get("photo")
        q=blood(photo=photo)
        q.save()
    return render(request,"Carousel/blood_carousel_img_form.html")

@login_required(login_url='/admin_login/')
def blood_carousel_img_table(request):
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        user_profile = None
    q = blood.objects.all()
    return render(request, "Carousel/blood_carousel_img_table.html", {"q": q, 'user_profile': user_profile})


def blood_carousel_img_edit(request,id):
    q=blood.objects.get(id=id)
    return render(request,"Carousel/blood_carousel_img_update.html",{'q':q})
    

def blood_carousel_img_update(request,id):
    if request.method =="POST":
        photo=request.FILES.get("photo")
        q=blood.objects.get(id=id)
        q.photo=photo;
        q.save()
        return redirect("/blood_carousel_img_table")
    return render(request,"Carousel/blood_carousel_img_update.html")

def blood_carousel_img_delete(request,id):
    q=blood.objects.get(id=id)
    q.delete();
    return redirect("/blood_carousel_img_table")



def basic_requires(request):
    b=basic_requirements.objects.all()
    return render(request,"basic_requirements.html",{'b':b})


def additional_requires(request):
    p=additional_requirements.objects.all()
    return render(request,"addtional_requirements.html",{'p':p})


def requestform(request):
    pt=patient_request.objects.all()
    if(request.method=="POST"):
        patient_name=request.POST['patient_name']
        contact_name=request.POST['contact_name']
        email=request.POST['email']
        gender=request.POST['gender']
        age=request.POST['age']
        mobilenumber=request.POST['mobilenumber']
        hospitalname=request.POST['hospitalname']
        address=request.POST['address']
        bloodgroup=request.POST['bloodgroup']
        bloodrequirements=request.POST['bloodrequirements']
        bloodtypes=request.POST['bloodtypes']
        pt=patient_request(patient_name=patient_name,contact_name=contact_name,email=email,gender=gender,age=age,mobilenumber=mobilenumber,hospitalname=hospitalname,address=address,bloodgroup=bloodgroup,bloodrequirements=bloodrequirements,bloodtypes=bloodtypes)
        pt.save()
        messages.success(request,"data posted")
    return render(request,"requestform.html")

def patientrequesttable(request):
    pt=patient_request.objects.all()
    return render(request,"patientrequesttable.html",{'pt':pt})


def patientrequestdata_delete(request,id):
    pt=patient_request.objects.get(id=id);
    pt.delete();
    return redirect("/patientrequesttable")






import requests
import random
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Donor , Address
from django.contrib.auth.hashers import make_password
from django.contrib.auth import logout



from django.http import JsonResponse

from cities_light.models import Country , Region , City , SubRegion

# Create your views here.

def get_states(request, country_id):
    states = Region.objects.filter(country_id=country_id)
    data = {'states': list(states.values('pk', 'name'))}
    return JsonResponse(data)

def get_districts(request, state_id):
    districts = SubRegion.objects.filter(region_id=state_id)
    data = {'districts': list(districts.values('pk', 'name'))}
    return JsonResponse(data)


def my_view(request):
    countries = Country.objects.all()
    
    return render(request, 'my_template.html', {'countries': countries})




def doner_home(request):
    return render (request,'doner-home.html')

def send_otp(number, message):
    url = "https://www.fast2sms.com/dev/bulk"
    api = "AQk4vziURB67Nsiq0fI6yad28gHMC6snwacnN0ZW5EDFr4lyuuYREzTGyJ8a"
    querystring = {"authorization": api, "sender_id": "FSTSMS", "message": message, "language": "english", "route": "p", "numbers": number}
    headers = {
        'cache-control': "no-cache"
    }
    return requests.request("GET", url, headers=headers, params=querystring)
def send_msg(number, message):
    url = "https://www.fast2sms.com/dev/bulk"
    api = "AQk4vziURB67Nsiq0fI6yad28gHMC6snwacnN0ZW5EDFr4lyuuYREzTGyJ8a"
    querystring = {"authorization": api, 
                   "sender_id": "FSTSMS", 
                   "message": message, 
                   "language": "english", 
                   "route": "p", 
                   "numbers": number}
    headers = {
        'cache-control': "no-cache"
    }
    return requests.request("GET", url, headers=headers, params=querystring)



from django.contrib.auth.models import User

def donor_register(request):
    countries = Country.objects.all()
    if request.method == "POST":
        fullname = request.POST.get('fullname')
        gender = request.POST.get('gender')
        age = request.POST.get('age')
        bloodgroup = request.POST.get('bloodgroup')
        aadhar = request.POST.get('aadhar')
        phone_number = request.POST.get('phone_number')
        email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        country_id = request.POST.get('country')
        state_id = request.POST.get('state')
        district_id = request.POST.get('district')

        address = Address.objects.create(
            country_id=country_id,
            state_id=state_id,
            district_id=district_id,
        )

        user = User(username=username, password=password, email=email)

        # Hash the password and save the User object
        user.set_password(password)
        user.save()

        donor = Donor.objects.create(
            fullname=fullname,
            gender=gender,
            age=age,
            bloodgroup=bloodgroup,
            aadhar=aadhar,
            phone_number=phone_number,
            email=email,
            address=address,
            user=user  # Link the Donor to the User
        )
        donor.save()

        messages.success(request, 'Registration successful')
        return redirect('/')

    return render(request, "doner-register.html", {'countries': countries})
from django.contrib.auth import authenticate, login
from  django.urls import reverse
def donor_login(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None and not (user.is_staff or user.is_superuser):
            login(request, user)
            
            referer = request.META.get('HTTP_REFERER')
            if referer:
                return redirect(referer)
            else:
                return redirect(reverse('home'))
        else:
            # Incorrect credentials - Return 401 Unauthorized status
            return JsonResponse({'message': 'Invalid Credentials'}, status=401)
    return JsonResponse({'message': 'Method Not Allowed'}, status=405)

def donor_logout(request):
    logout(request)
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    else:
        return redirect(reverse('home'))  


from django.core.exceptions import ObjectDoesNotExist 
def donor_forget_password(request):
    if request.method == "POST":
        phone_number = request.POST.get('phone_number')
        
        try:
            # Check if the provided phone number exists in the Donor database
            donor = Donor.objects.get(phone_number=phone_number)
            
            otp = random.randint(1000, 9999)
            otp_message = f'3 life 2 support: Your OTP for password reset is {otp}'
            send_msg(phone_number, otp_message)

            request.session['password_reset_data'] = {
                'phone_number': phone_number,
                'otp': otp,
            }

            return redirect('/password_reset_otp/')
        
        except ObjectDoesNotExist:
            messages.error(request, 'No account associated with this phone number')
                
    return render(request, "forget_password.html")


def password_reset_otp(request):
    if request.method == "POST":
        u_otp = request.POST.get('otp')
        password_reset_data = request.session.get('password_reset_data')

        if password_reset_data and u_otp.isdigit() and int(u_otp) == password_reset_data['otp']:
            request.session['reset_phone'] = password_reset_data['phone_number']
            return redirect('reset_password')
        else:
            messages.error(request, 'Wrong OTP')
    return render(request, "password_reset_otp.html")
def reset_password(request):
    if request.method == "POST":
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        reset_phone = request.session.get('reset_phone')

        if password == confirm_password and reset_phone:
            try:
                donor = Donor.objects.get(phone_number=reset_phone)
                hashed_password = make_password(password)
                donor.user.set_password(password)  # Use set_password to hash the password
                donor.user.save()  # Save the user object
                messages.success(request, 'Password reset successful')
                del request.session['password_reset_data']
                del request.session['reset_phone']
                return redirect('/')
            except Donor.DoesNotExist:
                pass
        else:
            messages.error(request, 'Passwords do not match or phone reset session expired')
            return JsonResponse({'message': 'Password reset failed'}, status=400)
    
    # Handle cases where the request is not a POST request (render the template)
    return render(request, "reset_password.html")

def bloodA(request):
    yes = bloodtype.objects.all()
    return render(request, 'bloodtypes.html', {'yes': yes } )

from .models import blood_types ,historyblood
def typesofblood(request):
    r=blood_types.objects.all()
    return render(request,"typesofblood.html",{'r':r})

def history(request):
    k=historyblood.objects.all()
    return render(request, 'historyofblood.html', {'k': k})


def donator(request):
    k=donators.objects.all()
    return render(request,"donators.html",{'k':k})

from .models import Announcements, Image ,video

def announce(request):
    a=Announcements.objects.all()
    return render(request,"announcement.html",{'a':a})

def image_slider(request):
    images = Image.objects.all()
    return render(request, 'image_slider.html', {'images': images})
def video1(request):
    v=video.objects.all()
    return render(request,"video.html",{'v':v})

def admindash(request):
    return render (request,'adminsite.html')

from django.http import HttpResponse
from .models import contactus


from .models import donation_process
def donation__process(request):
    if request.method=="GET":
        k=donation_process.objects.all()
    return render(request,"donation_process.html",{'k':k})
from .models import lastdonatedetails
from datetime import datetime
from django.utils import timezone
def lastdonform(request):
    if request.method == "POST":
        donor = request.user.donor
        hospital_name = request.POST['hospitalname']
        purpose = request.POST['purpose']
        patientname = request.POST['patientname']
        place = request.POST['address']
        date = request.POST['date']
        uploade_image = request.FILES.get("uploade_image")
        
        # Assuming `lastdonatedetails` is your model class
        q = lastdonatedetails(
            donor=donor,
            hospital_name=hospital_name,
            purpose=purpose,
            patient_name=patientname,
            donation_date=date, 
            place=place,
            uploade_image=uploade_image
        )
        
        # Save the data to the database
        q.save()

        # Update the last_donation_date field in the donor model
        donor.last_donation_date = datetime.strptime(date, '%Y-%m-%d').date()  # Assuming the date format is 'YYYY-MM-DD'
        
        today = timezone.now().date()
        next_donation_date = donor.last_donation_date + timezone.timedelta(days=90)  # Assuming a 90-day gap
        donor.save()

        donor.save()
        # Add a success message
        return redirect('/')
        return HttpResponse('Donation details saved successfully')
      
    

def last_don_retrieve(request):
    lastforms = lastdonatedetails.objects.all()
    saved_images = Image.objects.values_list('image_file', flat=True)
    return render(request , "lastdonrettrieve.html", {'lastforms': lastforms , 'saved_images':saved_images})
def save_image(request):
    if request.method == 'POST':
        selected_image_url = request.POST.get('selected_image')
        image_title = request.POST.get('image_title')
        
        # Create a new Image instance and save it to the database
        if selected_image_url and image_title:
            image = Image(title=image_title, image_file=selected_image_url)
            image.save()
            
            return HttpResponse('Image saved successfully')
    
    return redirect('last_don_retrieve')
def update_donors(request, donor_id):
    countries = Country.objects.all()
    donor = Donor.objects.get(id=donor_id)
    if request.method == "POST":
        fullname = request.POST.get('fullname')
        gender = request.POST.get('gender')
        age = request.POST.get('age')
        bloodgroup = request.POST.get('bloodgroup')
        aadhar = request.POST.get('aadhar')
        phone_number = request.POST.get('phone_number')
        email = request.POST.get('email')
        country_id = request.POST.get('country')
        state_id = request.POST.get('state')
        district_id = request.POST.get('district')
        address = Address.objects.create(
            country_id=country_id,
            state_id=state_id,
            district_id=district_id,
        )
        
        donor.fullname = fullname
        donor.gender = gender
        donor.age = age
        donor.bloodgroup = bloodgroup
        donor.aadhar = aadhar
        donor.phone_number = phone_number
        donor.email = email
        donor.address = address
        donor.save()
       
    return render(request, "update_donor_details.html", {'donor': donor,'countries':countries})


@login_required(login_url='/admin_login/')
def additional_table(request):
        try:
          user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
          user_profile = None
        p=additional_requirements.objects.all()
        return render(request,"additional_table.html",{'p':p,'user_profile':user_profile})


def additional_form(request):
    if request.method=="POST":
        # heading1=request.POST['heading1']
        image=request.FILES.get('image')
        point2=request.POST['point2']
        point3=request.POST['point3']
        point4=request.POST['point4']
        point5=request.POST['point5']
        point6=request.POST['point6']
        point7=request.POST['point7']
        point8=request.POST['point8']
        point9=request.POST['point9']
        point1=request.POST['point1']
        image1=request.FILES.get('image1')
        # image=request.FILES.get('image')
        heading=request.POST['heading']
        p=additional_requirements(image=image,point2=point2,point3=point3,point4=point4,point5=point5,point6=point6,point7=point7,point8=point8,point9=point9,
                               point1=point1,image1=image1,heading=heading)
        p.save()
        
        return redirect('/additional_table')
    return render(request,"additional_form.html")


def additional_edit(request,id):
    post=additional_requirements.objects.get(id=id)
    return render(request,"additional_update.html",{'post':post})



def additional_update(request,id):
    if request.method=="POST":
        image=request.FILES.get('image')
        point2=request.POST.get('point2')
        point3=request.POST.get('point3')
        point4=request.POST.get('point4')
        point5=request.POST.get('point5')
        point6=request.POST.get('point6')
        point7=request.POST.get('point7')
        point8=request.POST.get('point8')
        point9=request.POST.get('point9')
        point1=request.POST.get('point1')
        image1=request.FILES.get('image1')
        heading=request.POST.get('heading')
        post=additional_requirements.objects.get(id=id)
        post.image=image;
        post.point2=point2;
        post.point3=point3;
        post.point4=point4;
        post.point5=point5;
        post.point6=point6;
        post.point7=point7;
        post.point8=point8;
        post.point9=point9;
        post.point1=point1;
        post.image1=image1;
        post.heading=heading;
       
        post.save()
        return redirect("/additional_table")
    return render(request,"additional_update.html")



def additional_delete(request,id):
    q=additional_requirements.objects.get(id=id)
    q.delete();
    return redirect("/additional_table")

def eventform(request):
    k=events.objects.all()
    if request.method=="POST":
        content=request.POST['content']
        image=request.FILES.get('image')
        heading=request.POST['heading']
        k=events(content=content,imag=image,heading=heading)
        k.save()
     
     
    return render(request,"eventsform.html")

def event_edit(request,id):
    post=events.objects.get(id=id)
    return render(request,"events_update.html",{'post':post})

def event_update(request,id):
    if request.method =="POST":
        Content=request.POST.get('Content')
        
        Image=request.FILES.get('Image')
        Heading=request.POST.get('Heading')
        post=events.objects.get(id=id)
        post.content=Content;
        post.image=Image;
        post.heading=Heading;
       
        post.save()
        return redirect("/events_table")
    return render(request,"events_update.html")

def event_delete(request,id):
    q=events.objects.get(id=id)
    q.delete();
    return redirect("/events_table")

@login_required(login_url='/admin_login/')    
def event_table(request):
        try:
           user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
           user_profile = None
        post=events.objects.all()
        news_count = events.objects.count()  # Get the count of events
        return render(request, 'events_table.html', {"post":post,'news_count': news_count,'user_profile':user_profile})

  
def historyform(request):
    if request.method=="POST":
        heading=request.POST['heading']       
        images=request.FILES.get('images')
        points=request.POST['points']
        point1=request.POST['point1']
        point2=request.POST['point2']
        point3=request.POST['point3']
        point4=request.POST['point4']
        point5=request.POST['point5']
        subhead=request.POST['subhead']
        cnt1=request.POST['cnt1']
        cnt2=request.POST['cnt2']
        cnt3=request.POST['cnt3']
        cnt4=request.POST['cnt4']
        cnt5=request.POST['cnt5']
        b=historyblood(subhead=subhead,cnt5=cnt5,cnt4=cnt4,cnt3=cnt3,cnt2=cnt2,point5=point5,point4=point4,point3=point3,point2=point2,point1=point1,heading=heading,cnt1=cnt1,images=images,points=points)
        b.save()
        return redirect("/history_table")
        # return Httpresponse("data is inserted")
    return render(request,"historyform.html")
 
@login_required(login_url='/admin_login/')    
def history_table(request):
        try:
           user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
           user_profile = None
        h=historyblood.objects.all()
        return render(request,"history_table.html",{'h':h,'user_profile':user_profile})

def history_delete(request,id):
    c=historyblood.objects.get(id=id);
    c.delete();
    return redirect("/history_table");
    return render(request,'history_table.html',{'post':post})

def history_update(request,id):
   if(request.method=="POST"):
       Subhead=request.POST.get("Subhead");
       Heading=request.POST.get("Heading");
       Images=request.POST.get("Images");
       Point1=request.POST.get("Point1");
       Point2=request.POST.get("Point2");
       Point3=request.POST.get("Point3");
       Point4=request.POST.get("Point4");
       Points=request.POST.get("Points");
       Cnt1=request.POST.get("Cnt1");
       Cnt2=request.POST.get("Cnt2");
       Cnt3=request.POST.get("Cnt3");
       Cnt4=request.POST.get("Cnt4");
       Cnt5=request.POST.get("Cnt5");

       post=historyblood.objects.get(id=id);
       post.subhead=Subhead;
       post.heading=Heading;
       post.images=Images;
       post.point1=Point1;
       post.point2=Point2;
       post.point3=Point3;
       post.point4=Point4;
       post.points=Points;
       post.cnt1=Cnt1;
       post.cnt2=Cnt2;
       post.cnt3=Cnt3;
       post.cnt4=Cnt4;
       post.cnt5=Cnt5;


       post.save();
       return redirect("/history_table")
   return render(request,'history_update.html')

def history_edit(request,id):
    post=historyblood.objects.get(id=id);
    return render(request,'history_update.html',{'post':post})

def bloodtypesform(request):
    k=bloodtype.objects.all()
    if request.method=="POST":
        BId=request.POST['BId']
        Description=request.POST['Description']
        point1=request.POST['point1']
        point2=request.POST['point2']
        point3=request.POST['point3']
        point4=request.POST['point4']
        point5=request.POST['point5']
        point6=request.POST['point6']
        Bloodgroup=request.POST['Bloodgroup']

        image=request.FILES.get('image')
       
        k=bloodtype(BId=BId,Description=Description,point1=point1,point2=point2,point3=point3,image=image,point4=point4,point5=point5,point6=point6,Bloodgroup=Bloodgroup)
        k.save()
        return  redirect("/bloodtyperetrieve")
    return render(request,"bloodtypeform.html")
  
@login_required(login_url='/admin_login/')    
def bloodtyperetrieve(request):
        try:
          user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
          user_profile = None
        blood=bloodtype.objects.all()
        return render(request,"bloodtypesretrieve.html",{'blood':blood,'user_profile':user_profile})
    
def bloodtypesdelete(request,id):
    c=bloodtype.objects.get(id=id);
    c.delete();
    return redirect("/bloodtyperetrieve");
    return render(request,' bloodtypesretrieve.html',{'post':post})
def bloodtypesedit(request,id):
    post=bloodtype.objects.get(id=id)
    
    return render(request,'bloodtypeupdate.html',{'post':post})



def bloodtypesupdate(request,id):
   if(request.method=="POST"):
      bId=request.POST.get("bId");
      description=request.POST.get("description");
      Point1=request.POST.get("Point1");
      Point2=request.POST.get("Point2");
      Point3=request.POST.get("Point3");
      Point4=request.POST.get("Point4");
      Point5=request.POST.get("Point5");
      Point6=request.POST.get("Point6");
      bloodgroup=request.POST.get("bloodgroup");
      Image=request.FILES.get("Image");
   
      post=bloodtype.objects.get(id=id)
    
      post.BId=bId;
      post.Description=description;
      post.point1=Point1;
      post.point2=Point2;
      post.point3=Point3;
      post.point4=Point4;
      post.point5=Point5;
      post.point6=Point6;
      post.Bloodgroup=bloodgroup;
      post.image=Image;
      
      post.save();
      return redirect("/bloodtyperetrieve/")
   return render(request,'bloodtypeupdate.html')


def donorsform(request):
    k=donators.objects.all()
    if request.method=="POST":
        heading=request.POST['heading']
        Description=request.POST['Description']
       

        image=request.FILES.get('image')
       
        k=blood_types(heading=heading, Description=Description, image=image)
        k.save()
        return redirect('/donorsretrieve')
    return render(request,"donorform.html")

@login_required(login_url='/admin_login/')   
def donorsretrieve(request):
        try:
            user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
             user_profile = None
        blood=blood_types.objects.all()
        return render(request,"donorretrieve.html",{'blood':blood,'user_profile':user_profile})
    
def donorsdelete(request,id):
    c=blood_types.objects.get(id=id);
    c.delete();
    return redirect("/donorsretrieve");
    return render(request,' donorretrieve.html',{'post':post})

def donorsedit(request,id):
    post=blood_types.objects.get(id=id);
    return render(request,'donorupdate.html',{'post':post})
def donorsupdate(request,id):
   if(request.method=="POST"):
      
      Heading=request.POST.get("Heading");
     
      description=request.POST.get("description");
      
     
      Image=request.FILES.get("Image");
   
      post=blood_types.objects.get(id=id);
    
      post.heading=Heading;
      
      post.Description=description;
    
      post.image=Image;
     
      post.save();
      return redirect("/donorsretrieve")
   return render(request,'donorupdate.html')

def aboutus_form(request):
    k=aboutus.objects.all()
    if request.method=="POST":
        heading=request.POST['heading']
        subtitle1=request.POST['subtitle1']
        subtitle2=request.POST['subtitle2']
        content1=request.POST['content1']
        content2=request.POST['content2']
        content3=request.POST['content3']
        content4=request.POST['content4']
        content5=request.POST['content5']
        image1=request.FILES.get('image1')
        image2=request.FILES.get('image2')
        k=aboutus(content1=content1,content2=content2,content3=content3,content4=content4,content5=content5,image1=image1,image2=image2,heading=heading,subtitle1=subtitle1,subtitle2=subtitle2)
        k.save()

    return render(request,"aboutusform.html")
def aboutus_edit(request,id):
    post=aboutus.objects.get(id=id)
    return render(request,"aboutus_update.html",{'post':post})

def aboutus_update(request,id):
    if request.method =="POST":
        heading=request.POST['heading']
        subtitle1=request.POST['subtitle1']
        subtitle2=request.POST['subtitle2']
        content1=request.POST['content1']
        content2=request.POST['content2']
        content3=request.POST['content3']
        content4=request.POST['content4']
        content5=request.POST['content5']
        image1=request.FILES.get('image1')
        image2=request.FILES.get('image2')
        post=aboutus.objects.get(id=id)
        post.heading=heading;
        post.subtitle1=subtitle1;
        post.subtitle2=subtitle2;
        post.content1=content1;
        post.content2=content2;
        post.content3=content3;
        post.content4=content4;
        post.content5=content5;
        post.image1=image1;
        post.image2=image2;
        post.save()
        return redirect("/aboutus_table")
    return render(request,"aboutus_update.html")

def aboutus_delete(request,id):
    q=aboutus.objects.get(id=id)
    q.delete();
    return redirect("/aboutus_table")
    
@login_required(login_url='/admin_login/')    
def aboutus_table(request):
        try:
          user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
           user_profile = None
        post=aboutus.objects.all()
        return render(request,"aboutus_table.html",{"post":post,'user_profile':user_profile})



def insertphoto(request):
    if request.method=="POST":
        image_file= request.FILES.get('image_file')

        images=Image.objects.create(
            image_file=image_file
        )
        images.save()
        return redirect("/photo_table")
    return render(request,'upload_photo.html')

def photo_edit(request,id):
    images=Image.objects.get(id=id)
    return render(request,"photo_update.html",{'images':images})
    
def photo_update(request,id):
    if request.method =="POST":
        image_file=request.FILES.get("image_file")
        images=Image.objects.get(id=id)
        images.image_file=image_file;
        images.save()
        return redirect("/photo_table")
    return render(request,"photo_update.html")

def photo_delete(request,id):
    images=Image.objects.get(id=id)
    images.delete();
    return redirect("/photo_table")

@login_required(login_url='/admin_login/')
def photo_table(request):
        try:
          user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
          user_profile = None
        images=Image.objects.all()
        return render(request,"photo_table.html",{"images":images,'user_profile':user_profile})
    
def insertvideo(request):
    if request.method=="POST":
        video_url = request.POST['video_url'];
        v = video.objects.create(video_url=video_url)
        v.save()
        return redirect("/video_table")
    return render(request,"upload_video.html")
    
@login_required(login_url='/admin_login/')    
def video_table(request):
        try:
            user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
            user_profile = None
        q=video.objects.all()
        return render(request,"video_table.html",{"q":q,'user_profile':user_profile})


def video_edit(request,id):
    q=video.objects.get(id=id)
    return render(request,"video_update.html",{'q':q})
    

def video_update(request,id):
    if request.method =="POST":
        video_url=request.FILES.get("video_url")
        q=video.objects.get(id=id)
        q.video_url=video_url;
        q.save()
        return redirect("/video_table")
    return render(request,"video_update.html")

def video_delete(request,id):
    q=video.objects.get(id=id)
    q.delete();
    return redirect("/video_table")
     


def announcements_insert_form(request):
    if(request.method=="POST"):
        Title=request.POST['Title'];
        Content=request.POST['Content'];
        announcement=Announcements(Title=Title,Content=Content)
        announcement.save()
        return redirect("/announcements_retrive_table")
    return render(request,"announcements_insert_form.html")

@login_required(login_url='/admin_login/')
def announcements_retrive_table(request):
        try:
            user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
            user_profile = None
        a=Announcements.objects.all()
        return render(request,"announcements_retrive_table.html",{'a':a,'user_profile':user_profile})



def announcements_edit(request,id):
    a=Announcements.objects.get(id=id)
    return render(request,"announcements_update_form.html",{'a':a})

def announcements_update_form(request,id):
    if(request.method=="POST"):
        Title=request.POST.get("Title")
        Content=request.POST.get("Content")
        a=Announcements.objects.get(id=id)
        a.Title=Title;
        a.Content=Content;
        a.save()
        return redirect("/announcements_retrive_table")
    return render(request,"announcements_update_form.html")

def announcements_delete(request,id):
    a=Announcements.objects.get(id=id)
    a.delete()
    return redirect("/announcements_retrive_table")


def contactform(request):
    if request.method =="POST":
        heading=request.POST["heading"]
        address=request.POST["address"]
        email = request.POST["email"]
        
        bloodtype = contact.objects.create(
            heading=heading,
            address=address,
            email=email,
        )
        bloodtype.save()

        return redirect('/contactretrive')
    return render(request, "contactinsert.html")



def contactupdate(request, id):
     
    
    if request.method == "POST":
        Heading = request.POST.get("Heading")
        Address = request.POST.get("Address")
        Email = request.POST.get("Email")
        post=contact.objects.get(id=id)
        post.heading = Heading
        post.address = Address
        post.email = Email
        post.save()
        return redirect("/contactretrive")
    
    return render(request, "contactupdate.html")

@login_required(login_url='/admin_login/')
def contactretrive(request):
    post = contact.objects.all()
    return render(request, "contactretre.html", {'post': post})
     


def contactedit(request, id):
    post = contact.objects.get(id=id)
    return render(request, "contactupdate.html", {'post': post})

def contactdelete(request, id):
    post = contact.objects.get(id=id)
    post.delete()
    return redirect("/contactretrive")

@login_required(login_url='/admin_login/')
def registerretrive(request):
        try:
           user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
           user_profile = None
        donor_data = Donor.objects.all()    
        return render(request, "registerdlist.html", {'donor_data': donor_data,'user_profile':user_profile})



def hidedonor(request,id):
    d=get_object_or_404(Donor,pk=id)
    if request.method == "POST":
        in_hidden=request.POST['in_hidden']
        d.in_hidden=in_hidden
        d.save()
        return redirect("/registerretrive/")
    return render(request,"hide.html",{'d':d})

def showdonor(request,id):
    d=get_object_or_404(Donor,pk=id)
    if request.method == "POST":
        in_hidden=request.POST['in_hidden']
        d.in_hidden=in_hidden
        d.save()
        return redirect("/registerretrive/")
    return render(request,"showdonor.html",{'d':d})


def registerdelete(request, id):
    donor_data = Donor.objects.get(id=id)
    donor_data.delete()
    return redirect('/registerretrive')

@login_required(login_url='/admin_login/')    
def footer_table(request):
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        user_profile = None
    post=socialicon.objects.all()
    p=appfooter.objects.all()
    p1=appfooter1.objects.all()
    p2=appdown.objects.all()
    p3=gettouch.objects.all()
    return render(request,"footer_crud/footer_table.html",{'post':post,'p1':p1,'p2':p2,'p3':p3,'p':p,'user_profile':user_profile})

def donationform(request):
    if request.method =="POST":
        heading=request.POST["heading"]
        point1=request.POST["point1"]
        point2 = request.POST["point2"]
        point3 = request.POST["point3"]

        point4 = request.POST["point4"]

       
        image=request.FILES.get('image')

        
        b = donation_process.objects.create(
            heading=heading,
            point1=point1,
            point2=point2,
            point3=point3,
            point4=point4,
          
            image=image
        )
        b.save()

        return redirect('/donationprocess_table')
    return render(request, "donationprocessform.html")

@login_required(login_url='/admin_login/')    
def donationprocess_table(request):
        try:
           user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
            user_profile = None
        blood=donation_process.objects.all
        return render(request,"donationprocess_table.html",{'blood': blood,'user_profile':user_profile})
    
def donationprocess_delete(request,id):
    c=donation_process.objects.get(id=id);
    c.delete();
    return redirect("/donationprocess_table");
    return render(request,' donationprocess_table.html',{'post':post})

def donationprocess_update(request,id):
   if(request.method=="POST"):
      
       Heading=request.POST.get("Heading");
      
       Point1=request.POST.get("Point1");
       Point2=request.POST.get("Point2");
       Point3=request.POST.get("Point3");
       Point4=request.POST.get("Point4");
       Point5=request.POST.get("Point5");
       Point6=request.POST.get("Point6");
       
       Image=request.FILES.get("Image");
   
       post=donation_process.objects.get(id=id);
    
       post.heading=Heading;
     
       post.point1=Point1;
       post.point2=Point2;
       post.point3=Point3;
       post.point4=Point4;
       post.point5=Point5;
       post.point6=Point6;
       post.image=Image;
      
       post.save();
       return redirect("/donationprocess_table")
   return render(request,'donationprocess_update.html')

def donationprocess_edit(request,id):
    post=donation_process.objects.get(id=id);
    return render(request,'donationprocess_update.html',{'post':post})

@login_required(login_url='/admin_login/')
def tips_table(request):
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        user_profile = None
    post=tips_donation.objects.all()
    c=content.objects.all()
    return render(request,"tips_table.html",{"post":post,'c':c,'user_profile':user_profile})


def tips_form(request):
    if request.method=="POST":
         # heading1=request.POST['heading1']
        subHeading=request.POST['subHeading']
        point1=request.POST['point1']
        point2=request.POST['point2']
        point3=request.POST['point3']
        t=tips_donation(subHeading=subHeading,point1=point1,point2=point2,point3=point3)
        t.save()
        
        return redirect("/tips_table")
    return render(request,'tipsform.html')
def tips_edit(request,id):
    post=tips_donation.objects.get(id=id)
    return render(request,"tips_update.html",{'post':post})


def content_form(request):
     if request.method=='POST':
        img1=request.FILES.get('img1')
        Heading1=request.POST['Heading1']
        description=request.POST['description']
        text=request.POST['text']
        point1=request.POST['point1']
        point2=request.POST['point2']
        point3=request.POST['point3']
        c=content(img1=img1,Heading1=Heading1,description=description,text=text,point1= point1,point2=point2,point3=point3)
        c.save()
        return redirect("/tips_table")
     return render(request,"contentform.html")

def tips_update(request,id):
    if request.method=="POST":
        SubHeading=request.POST.get('SubHeading')
        Point1=request.POST.get('Point1')

        Point2=request.POST.get('Point2')
        Point3=request.POST.get('Point3')
        post=tips_donation.objects.get(id=id)
        post.subHeading=SubHeading;
        post.point1=Point1;
        post.point2=Point2;
        post.point3=Point3;
       
        post.save()
        return redirect("/tips_table")
    return render(request,"tips_update.html")

def tips_delete(request,id):
    q=tips_donation.objects.get(id=id)
    q.delete();
    return redirect("/tips_table")

def content_update(request,id):
    if request.method=="POST":
        img1=request.POST.get('img1')
        Heading1=request.POST.get('Heading1')
        description=request.POST.get('description')
        text=request.POST.get('text')
        
        point1=request.POST.get('point1')
        point2=request.POST.get('point2')
        point3=request.POST.get('point3')
        post=content.objects.get(id=id)
        post.img1=img1
        post.Heading1=Heading1
        post.text=text
        post.description=description;
        post.point1=point1;
        post.point2=point2;
        post.point3=point3;
        post.save()
        return redirect("/tips_table")
    return render(request,"content_update.html")

def content_delete(request,id):
    q=content.objects.get(id=id)
    q.delete();
    return redirect("/tips_table")

def content_edit(request,id):
    post=content.objects.get(id=id)
    return render(request,"content_update.html",{'post':post})
def basic_requires(request):
    b=basic_requirements.objects.all()
    return render(request,"basic_requirements.html",{'b':b})

@login_required(login_url='/admin_login/')   
def basic_table(request):
        try:
          user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
          user_profile = None
        blood=basic_requirements.objects.all()
        return render(request,"basic_table.html",{'blood':blood,'user_profile':user_profile})
    
def basicform(request):
    if request.method=="POST":
        heading=request.POST['heading']
        icon=request.FILES.get('icon')
        images=request.FILES.get('images')
        points=request.POST['points']
        points=request.POST['points']

       
        point1=request.POST['point1']

        point2=request.POST['point2']

        point3=request.POST['point3']

        point4=request.POST['point4']

        point5=request.POST['point5']

        point6=request.POST['point6']

        point7=request.POST['point7']

        point8=request.POST['point8']

        point9=request.POST['point9']

       


        b=basic_requirements(point9=point9,point8=point8,point7=point7,point6=point6,point5=point5,point4=point4,point3=point3,point2=point2,point1=point1,heading=heading,icon=icon,images=images,points=points)
        b.save()
        return redirect("/basic_table")
        # return Httpresponse("data is inserted")
    return render(request,"basicform.html")
def basic_delete(request,id):
    c=basic_requirements.objects.get(id=id);
    c.delete();
    return redirect("/basic_table");
    return render(request,' donationprocess_table.html',{'post':post})

def basic_update(request,id):
   if(request.method=="POST"):
       Heading=request.POST.get("Heading");
       Icon=request.POST.get("Icon");
       Images=request.POST.get("Images");
       Points=request.POST.get("Points");
       Points=request.POST.get("Points");

       Point1=request.POST.get("Point1");
       Point2=request.POST.get("Point2");
       Point3=request.POST.get("Point3");
       Point4=request.POST.get("Point4");
       Point5=request.POST.get("Point5");
       Point6=request.POST.get("Point6");
       Point7=request.POST.get("Point7");
       Point8=request.POST.get("Point8");
       Point9=request.POST.get("Point9");

     
   
       post=basic_requirements.objects.get(id=id);
       post.heading=Heading;
       post.icon=Icon;
       post.images=Images;
       post.points=Points;
       post.points=Points;


       post.point1=Point1;
       post.point2=Point2;
       post.point3=Point3;
       post.point4=Point4;
       post.point5=Point5;
       post.point6=Point6;
       post.point7=Point7;

       post.point8=Point8;

       post.point9=Point9;



      
       post.save();
       return redirect("/basic_table")
   return render(request,'basic_update.html')

def basic_edit(request,id):
    post=basic_requirements.objects.get(id=id);
    return render(request,'basic_update.html',{'post':post})
def contactus_form(request):
    k=Contactusadmin.objects.all()
    if request.method=="POST":
        email=request.POST['email']
        contactno=request.POST['contactno']
        website=request.POST['website']
        address=request.POST['address']
        map=request.POST['map']
        
        k=Contactusadmin(email=email,contactno=contactno,website=website,address=address,map=map)
        k.save()
    return render(request,"contactus_form.html")
def contactus_edit(request,id):
    post=Contactusadmin.objects.get(id=id)
    return render(request,"contactus_update.html",{'post':post})

def contactus_update(request,id):
    if request.method =="POST":
        email=request.POST['email']
        contactno=request.POST['contactno']
        website=request.POST['website']
        address=request.POST['address']
        map=request.POST['map']
        post=Contactusadmin.objects.get(id=id)
        post.email=email;
        post.contactno=contactno;
        post.website=website;
        post.address=address;
        post.map=map;
        post.save()
        return redirect("/contactus_table")
    return render(request,"contactus_update.html")

def contactus_delete(request,id):
    q=Contactusadmin.objects.get(id=id)
    q.delete();
    return redirect("/contactus_table")
    

def contactus_table(request):
        try:
          user_profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
          user_profile = None
        post=Contactusadmin.objects.all()
        a=contactus.objects.all()
        return render(request,"contactus_table.html",{"post":post,"a":a,'user_profile':user_profile})
    
def contactusadmin(request):
    if request.method=="GET":
        k=Contactusadmin.objects.all()
        return render(request,"contactusadmin.html",{'k':k})
def contactform(request):
    d=Contactusadmin.objects.all()
    if request.method =="POST":
        name=request.POST["name"]
        contact=request.POST["contact"]
        city = request.POST["city"]
        email = request.POST["email"]
        enquirytype= request.POST["enquirytype"]
        subject = request.POST["subject"]
        message = request.POST["message"]
        
        
        bloodtype = contactus.objects.create(
            name=name,
            city=city,
            email=email,
            enquirytype=enquirytype,
            subject=subject,
            message=message,
            contact=contact,
            
        )
        bloodtype.save()

   

 
    return render(request, "contactus.html",{'d':d})

def contactdelete(request,id):
    q=contactus.objects.get(id=id)
    q.delete();
    return redirect("/contactretrieve")
def contactus_form(request):
    k=Contactusadmin.objects.all()
    if request.method=="POST":
        email=request.POST['email']
        contactno=request.POST['contactno']
        website=request.POST['website']
        address=request.POST['address']
        map=request.POST['map']
        
        k=Contactusadmin(email=email,contactno=contactno,website=website,address=address,map=map)
        k.save()
    return render(request,"contactus_form.html")
def contactus_edit(request,id):
    post=Contactusadmin.objects.get(id=id)
    return render(request,"contactus_update.html",{'post':post})

def contactus_update(request,id):
    if request.method =="POST":
        email=request.POST['email']
        contactno=request.POST['contactno']
        website=request.POST['website']
        address=request.POST['address']
        map=request.POST['map']
        post=Contactusadmin.objects.get(id=id)
        post.email=email;
        post.contactno=contactno;
        post.website=website;
        post.address=address;
        post.map=map;
        post.save()
        return redirect("/contactus_table")
    return render(request,"contactus_update.html")

def contactus_delete(request,id):
    q=Contactusadmin.objects.get(id=id)
    q.delete();
    return redirect("/contactus_table")
def contactus_table(request):
    post=Contactusadmin.objects.all()
    a=contactus.objects.all()
    return render(request,"contactus_table.html",{"post":post,"a":a})
def contactusadmin(request):
    if request.method=="GET":
        k=Contactusadmin.objects.all()
        return render(request,"contactusadmin.html",{'k':k})
def contactform(request):
    d=Contactusadmin.objects.all()
    if request.method =="POST":
        name=request.POST["name"]
        contact=request.POST["contact"]
        city = request.POST["city"]
        email = request.POST["email"]
        enquirytype= request.POST["enquirytype"]
        subject = request.POST["subject"]
        message = request.POST["message"]
        
        
        bloodtype = contactus.objects.create(
            name=name,
            city=city,
            email=email,
            enquirytype=enquirytype,
            subject=subject,
            message=message,
            contact=contact,
            
        )
        bloodtype.save()

   

 
    return render(request, "contactus.html",{'d':d})

def contactdelete(request,id):
    q=contactus.objects.get(id=id)
    q.delete();
    return redirect("/contactretrieve")

import openpyxl
from django.http import HttpResponse

def table_to_excel(request):
    # Get the data from the Donor model
    donors = Donor.objects.all()

    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add the table headers to the first row of the worksheet
    headers = ['Full Name', 'Age', 'Gender', 'Bloodgroup', 'Aadhar No','Phone No','Email id','Country','State','district']
    for i, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=i, value=header)

    # Add the donor data to the worksheet
    for i, donor in enumerate(donors, start=2):
        worksheet.cell(row=i, column=1, value=donor.fullname)
        worksheet.cell(row=i, column=2, value=donor.age)
        worksheet.cell(row=i, column=3, value=donor.gender)
        worksheet.cell(row=i, column=4, value=donor.bloodgroup)
        worksheet.cell(row=i, column=5, value=donor.aadhar)
        worksheet.cell(row=i, column=6, value=donor.phone_number)
        worksheet.cell(row=i, column=7, value=donor.email)
        worksheet.cell(row=i, column=8, value=donor.address.country.name)
        worksheet.cell(row=i, column=9, value=donor.address.state.name)
        worksheet.cell(row=i, column=10, value=donor.address.district.name)

    # Set the filename and content type of the response
    filename = 'list of donors.xlsx'
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # Create the response object and set the headers
    response = HttpResponse(content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response and return it
    workbook.save(response)
    return response
def pre_registation_page(request):
    return render(request,"Questions/pre_reg_page.html")

def form_two_redirct(request):
    if request.method == 'POST':
        selected_option = request.POST.get('question_1')

        if selected_option == 'yes':
            return redirect('pre_reg_QA_2')  # Assuming 'base_template' is the URL name for the template
        elif selected_option == 'no':
            # Handle the case for 'no' option, perhaps redirect to another template
            return redirect('check_age') 

            

def check_age(request):
    return render(request,"Questions/check_age_no.html")

def form_three_redirct(request):
    if request.method == 'POST':
        selected_option = request.POST.get('question_2')

        if selected_option == 'yes':
            return redirect('check_weight')# Assuming 'base_template' is the URL name for the template
        elif selected_option == 'no':
            # Handle the case for 'no' option, perhaps redirect to another template
            return redirect('pre_reg_QA_3') 

            
def check_weight(request):
    return render(request,"Questions/check_weight_no.html")

def form_four_redirct(request):
    if request.method == 'POST':
        selected_option = request.POST.get('question_3')

        if selected_option == 'yes':
             return redirect('check_timeduration') # Assuming 'base_template' is the URL name for the template
        elif selected_option == 'no':
            # Handle the case for 'no' option, perhaps redirect to another template
            return redirect('pre_reg_QA_4')

             

def check_timeduration(request):
    return render(request,"Questions/check_timeduration_no.html")


def form_reg_page_redirect(request):
    if request.method == 'POST':
        selected_option = request.POST.get('question_4')

        if selected_option == 'yes':
            return redirect('check_heartprob') # Assuming 'base_template' is the URL name for the template
        elif selected_option == 'no':
            # Handle the case for 'no' option, perhaps redirect to another template
             return redirect('reg_page') 

            

def check_heartprob(request):
    return render(request,"Questions/check_heartprob_no.html")



def pre_reg_QA_1(request):
    if request.method=="post":
        question_1=request.post["question_1"]
        obj1=pre_registataion_QA_1(question_1=question_1)
        obj1.save()
    return render(request,"Questions/pre_reg_QA1.html")

def pre_reg_QA_2(request):
    if request.method=="post":
        question_2=request.post["question_2"]
        obj2=pre_registataion_QA_2(question_2=question_2)
        obj2.save()
    return render(request,"Questions/pre_reg_QA2.html")

def pre_reg_QA_3(request):
    if request.method=="post":
        question_3=request.post["question_3"]
        obj3=pre_registataion_QA_3(question_3=question_3)
        obj3.save()
    return render(request,"Questions/pre_reg_QA3.html")

def pre_reg_QA_4(request):
    if request.method=="post":
        question_4=request.post["question_4"]
        obj4=pre_registataion_QA_4(question_4=question_4)
        obj4.save()
    return render(request,"Questions/pre_reg_QA4.html")

def reg_page(request):
    return render(request,"Questions/registations_page.html")

def basic_requires(request):
    b=basic_requirements.objects.all()
    return render(request,"basic_requirements.html",{'b':b})

def additional_requires(request):
    p=additional_requirements.objects.all()
    return render(request,"addtional_requirements.html",{'p':p})
 
@login_required(login_url='/admin_login/')   
def nav_list(request):
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        user_profile = None
    nav_items = HomeNav_drop.objects.all()
    return render(request, 'nav_bar_crud/nav_list.html', {'nav_items': nav_items,'user_profile':user_profile})

def nav_form(request):
    k=HomeNav_drop.objects.all()
    if request.method=="POST":
        nav_name=request.POST['nav_name']
        image=request.FILES.get('image')
        nav_url=request.POST['nav_url']
        order=request.POST['order']
        parent_category=request.POST[parent_category]


        k=HomeNav_drop(nav_name=nav_name,imag=image,nav_url=nav_url,order=order,parent_category=parent_category)
        k.save()
     
     
    return render(request,"eventsform.html")




    
    
def read(request,id):
        d=get_object_or_404(UserProfile,pk=id)
        if request.method == "POST":
            aread= request.POST["aread"]
            d.aread=aread
            d.save()
            return redirect("/search_user/")
        return render(request,"areadsuc.html",{'d':d})

def readhide(request,id):
        d=get_object_or_404(UserProfile,pk=id)
        if request.method == "POST":
            aread= request.POST["aread"]
            d.aread=aread
            d.save()
            return redirect("/search_user/")
        return render(request,"homeahide.html",{'d':d})

def evebloodshow(request,id):
        d=get_object_or_404(UserProfile,pk=id)
        if request.method == "POST":
            awrite= request.POST["awrite"]
            d.awrite=awrite
            d.save()
            return redirect("/search_user/")
        return render(request,"eveblood.html",{'d':d})

def evebloodhide1(request,id):
        d=get_object_or_404(UserProfile,pk=id)
        if request.method == "POST":
            awrite= request.POST["awrite"]             
            d.awrite=awrite
            d.save()
            return redirect("/search_user/")
        return render(request,"eveblodhide.html",{'d':d})

def bloodlistshow(request,id):
        d=get_object_or_404(UserProfile,pk=id)
        if request.method == "POST":
            aupdate= request.POST["aupdate"]
            d.aupdate=aupdate
            d.save()
            return redirect("/search_user/")
        return render(request,"donorlist.html",{'d':d})

def bloodlisthide(request,id):
        d=get_object_or_404(UserProfile,pk=id)
        if request.method == "POST":
            aupdate= request.POST["aupdate"]             
            d.aupdate=aupdate
            d.save()
            return redirect("/search_user/")
        return render(request,"donorlisthide.html",{'d':d})


def gallashow(request,id):
        d=get_object_or_404(UserProfile,pk=id)
        if request.method == "POST":
            aremove= request.POST["aremove"]
            d.aremove=aremove
            d.save()
            return redirect("/search_user/")
        return render(request,"gallshow.html",{'d':d})

def gallahide(request,id):
        d=get_object_or_404(UserProfile,pk=id)
        if request.method == "POST":
            aremove= request.POST["aremove"]             
            d.aremove=aremove
            d.save()
            return redirect("/search_user/")
        return render(request,"gallhide.html",{'d':d})

def contshow(request,id):
        d=get_object_or_404(UserProfile,pk=id)
        if request.method == "POST":
            aconta= request.POST["aconta"]
            d.aconta=aconta
            d.save()
            return redirect("/search_user/")
        return render(request,"contashow.html",{'d':d})

def conthide(request,id):
        d=get_object_or_404(UserProfile,pk=id)
        if request.method == "POST":
            aconta= request.POST["aconta"]             
            d.aconta=aconta
            d.save()
            return redirect("/search_user/")
        return render(request,"contahide.html",{'d':d})




def search_user(request):
       countries = Country.objects.all()
      # Fetch your donor list based on your criteria
       userlist = User.objects.filter(is_superuser=False,is_staff=1).select_related('userprofile') 
       if request.method=="POST":       
        country_id = request.POST.get('country')
        state_id = request.POST.get('state')
        district_id = request.POST.get('district')     
        if request.user.is_authenticated:
            userlist = User.objects.filter(is_superuser=False,is_staff=1).select_related('userprofile') 
            if country_id:
                userlist = userlist.filter(address__country_id=country_id)
            if state_id:
                userlist = userlist.filter(address__state_id=state_id)
            if district_id:
                userlist = userlist.filter(address__district_id=district_id)
       return render(request,"address.html" ,{'userlist':userlist , 'countries' :countries})


def admin_register(request):
    countries = Country.objects.all()
    error_message = ''
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email')
        is_staff=request.POST.get('is_staff')
        country= request.POST.get('country')
        state_id = request.POST.get('state')
        district_id = request.POST.get('district')
        
        if username and password:
            if User.objects.filter(username=username).exists():
                error_message = "Username already taken. Please choose another username."
            else:
                user = User.objects.create_user(username=username,email=email, password=password,is_staff=is_staff )                
                user.save()
                userdetailes=UserProfile.objects.create( user=user)
                userdetailes.aread=request.POST['aread']
                userdetailes.awrite=request.POST['awrite']
                userdetailes.aupdate=request.POST['aupdate']
                userdetailes.aremove=request.POST['aremove']
                userdetailes.aconta=request.POST['aconta']
                userdetailes.mobileno=request.POST['mobileno']
                userdetailes.address= Address.objects.create(
                country_id=country,
                state_id=state_id,
                district_id=district_id,            
        )
                userdetailes.save()
                success_message = "Registration successful. You can now log in."
                return redirect('/admin_login')  
            
        else:
            error_message = "Both username and password are required."

    return render(request, 'registration/registration_form.html', {'error_message': error_message,'countries':countries})


def admin_login(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['pwd']
        user = authenticate(username=username, password=password)
        if user is not None and (user.is_superuser or user.is_staff):
            login(request, user)  
            return redirect('/administrate/')
        else:
            return HttpResponse('Check username and password')
    return render(request, 'registration/login.html')

def logout_admin(request):
    logout(request)
    return redirect('/admin_login/')
    
from django.contrib.auth import  update_session_auth_hash 
@login_required(login_url='/admin_login/')
def change_password(request):
    if request.method == 'POST':
        old_password = request.POST['old_password']
        new_password = request.POST['new_password']
        confirm_new_password = request.POST['confirm_new_password']

        if new_password != confirm_new_password:
            messages.error(request, 'New password and confirm password do not match')
        elif request.user.check_password(old_password):
            request.user.set_password(new_password)
            request.user.save()
            update_session_auth_hash(request, request.user)  
            messages.success(request, 'Password changed successfully')
            return redirect('/admin_login/')
        else:
            messages.error(request, 'Incorrect old password')

    return render(request, 'registration/change_password.html')   
    
def update_donors(request, donor_id):
    countries = Country.objects.all()
    donor = Donor.objects.get(id=donor_id)
    if request.method == "POST":
        fullname = request.POST.get('fullname')
        gender = request.POST.get('gender')
        age = request.POST.get('age')
        bloodgroup = request.POST.get('bloodgroup')
        aadhar = request.POST.get('aadhar')
        phone_number = request.POST.get('phone_number')
        email = request.POST.get('email')
        country_id = request.POST.get('country')
        state_id = request.POST.get('state')
        district_id = request.POST.get('district')
        address = Address.objects.create(
            country_id=country_id,
            state_id=state_id,
            district_id=district_id,
        )
        preferences = request.POST.get('preferences')
        donor.fullname = fullname
        donor.gender = gender
        donor.age = age
        donor.bloodgroup = bloodgroup
        donor.aadhar = aadhar
        donor.phone_number = phone_number
        donor.email = email
        donor.address = address
        donor.preferences = preferences
        donor.save()
        return redirect('/')

    return render(request, "update_donor_details.html", {'donor': donor,'countries':countries})
                
        

def donor_details(request, donor_id):
    donor = Donor.objects.get(id=donor_id)
    address = Address.objects.get(id=donor.address.id)
    return render(request, "donor_profile.html", {'donor': donor, 'address': address})
    
    
def search_user(request):
    if request.user.is_superuser:
     countries = Country.objects.all()
     userlist = User.objects.filter(is_superuser=False, is_staff=1).select_related('userprofile__address__district')

     if request.method == "POST":
        country_id = request.POST.get('country')
        state_id = request.POST.get('state')
        district_id = request.POST.get('district')

        # Create an empty Q object to build the filter conditions
        filters = Q()

        if country_id:
            filters &= Q(userprofile__address__country_id=country_id)
        if state_id:
            filters &= Q(userprofile__address__state_id=state_id)
        if district_id:
            filters &= Q(userprofile__address__district_id=district_id)

        # Apply the filters to the userlist
        userlist = userlist.filter(filters)
     return render(request, "address.html", {'userlist': userlist, 'countries': countries})
    else:
        return HttpResponseForbidden("You do not have permission to access this page.") 
        
def nav_update(request, pk):
    nav_item = get_object_or_404(HomeNav_drop, pk=pk)
    
    if request.method == 'POST':
        nav_item.nav_name = request.POST['nav_name']
        nav_item.nav_url = request.POST['nav_url']
        nav_item.parent_category_id = request.POST.get('parent_category')
        nav_item.order = request.POST['order']
        nav_item.save()
        return redirect('nav_list')
        
    return render(request, 'nav_bar_crud/nav_form.html', {'nav_item': nav_item})

def nav_delete(request, pk):
    nav_item = get_object_or_404(HomeNav_drop, pk=pk)
    
    if request.method == 'POST':
        nav_item.delete()
        return redirect('nav_list')
        
    return render(request, 'nav_bar_crud/nav_confirm_delete.html', {'nav_item': nav_item})

